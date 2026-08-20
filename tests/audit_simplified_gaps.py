"""Read-only audit of residual simplified Chinese in BOM business fields."""
from __future__ import annotations

import importlib.util
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl


BASE = Path(r"E:\OneDrive\Work Database\規格與認證\EC Tracking\每日暫存區")
CMS_PATH = BASE / "資料轉換" / "convert_multi_sheets.py"

spec = importlib.util.spec_from_file_location("cms", CMS_PATH)
cms = importlib.util.module_from_spec(spec)
assert spec.loader
spec.loader.exec_module(cms)

# Characters whose simplified form differs from the Taiwan traditional form.
# This is only a detector. Any addition to the business-rule mapping still
# requires a context review and user approval.
SIMPLIFIED_ONLY = set(
    "万与专业丛东丝丢两严丧个丰临为丽举么义乌乐乔习乡书买乱争于亏云亚产亩亲亵亿仅从仓仪们价众优会伞伟传伤伦伪体余佣佥侠侣侥侦侧侨俩俭债倾偿儿兑党兰关兴养兽冈册写军农冯冲决况冻净凄凉减凑凛几凤凭凯击凿刍划刘则刚创删别刬剂剐剑剥剧劝办务动励劲劳势匀区医华协单卖卢卤卧卫却厂厅历厉压厌厕厘厢厦厨县叁参双发变叙叠叶号叹吁吃合吊同后向吗吨听启吴呐呒呓呕员呛呜咏咙咛咝咤咨咸响哑哒哓哔哗哙哜哝哟唇唤啧啬啭啮喂喷喽嗫嘘嘤噜嚣团园围国图圆圣场坏块坚坛坝坞坟坠垄垅垆垒垫垦垩垱垲埘埙埚堑堕墙壮声壳处备复够头夸夹夺奁奋奖奥妆妇妈妩娱娄娇娱婴婵孙学宁宝实宠审宪宫宽宾寝对寻导寿将尔尘尝尧尸尽层屉屡属岗岘岛岭岳岿峄峡峣峤峥峦币帅师帐帘帜带帮帱帻帼幂干并广庄庆庐库应庙庞废开异弃张弥弯弹强归当录彦彻径徕忆忧忾怀态怂怃怄怅怆怜总恋恳恶恸恹恺恻恼恽悦悫悬悭惊惧惨惩惫惬惭惮惯愠愤愿慑懑懒戆戏战户执扩扫扬扰抚抛抢护报担拟拢拣拥拦拧拨择挂挚挛挞挟挠挡挣挤挥捞损捡换捣据掳掷掸掺揽搀搁搂搅摄摆摇摈摊撄撑撵撷撸撺擞攒敌敛数斋斗斩断无旧时旷昙显晋晒晓晔晕暂术朴机杀杂权条来杨杰极构枢枣枪枫柜柠查柽栀栅标栈栋栏树样桠桡桢档桥桦桧桨梦梼检棂椁椟椠椤椭楼榄榅榈榉槚槛槟横樯橱橹橼檩欢欧殁殇残殒殓殚殡殴毁毂毕毙气汇汉汤沟没沣沤沥沦沧沪泞泪泶泷泸泺泻泽洁洒浅浆浇浈浊测济浏浐浑浒浓浔浕涂涌涛涝涞涟涠涡涣涤润涧涨涩淀渊渌渍渎渐渔渗温湾湿溃溅滚滞满滤滥滨滩潆潇潋潍潜潴澜濑濒灭灯灵灾灿炀炉炖炜炝点炼炽烁烂烛烟烦烧烨烩烬热焕焖焘爱爷牍牵牺犊状犷犹狈狝狞独狭狮狯狰狱狲猃猎猕猪猫献獭玑玛玮环现玱玺珐珑琐琼瑶璎瓒电画畅畴疖疗疟疠疡疮疯痈痉痒痨痪瘘瘫瘾瘿癞癣皑皱盏盐监盖盘眍眦睁睐睑瞒矫矿砀码砖砗砚砺础硁硅硕硖硗硙确硷碍碛碜碱礼祢祯祷祸禀禄离秃秆积称秽稳窃窍窎窑窜窝窥窦竖竞笃笋笔笕笺笼笾筑筛筝筹签简箓箦箧箨箩箪箫篑篓篮篱簖籁类籴粜粤粪粮紧纠红纤约级纨纩纪纫纬纭纯纰纱纲纳纵纶纷纸纹纺纽线绀练组绅细织终绉绊绍绎经绐绑绒结绕绘给绚络绝绞统绠绡绢绣绤绥绦继绩绪续绰绳维绵绷绸综绽绿缀缁缂缃缄缅缆缇缈缉缊缋缌缎缏缑缒缓缔缕编缗缘缙缚缛缜缝缟缠缡缢缣缤缥缦缧缨缩缪缫缬缭缮缯缰缱缴缵罂网罗罚罢羁翘耸耻聂聋职联聪肃肠肤肮肴肾肿胀胁胆胜胧胪胫胶脉脍脏脐脑脓脔脚脱脸腊腌腘腭腻腾膑臜舆舰舱艰艳艺节芈芗芜芦苁苇苈苋苌苍苎苏苹范茎茏茑茔茧荆荐荙荚荛荜荞荟荠荡荣荤荥荦荧荨荩荪荫药莱莲莳获莸莹莺莼萚萝萤营萦萧萨葱蒇蒉蒋蒌蓝蓟蓠蓣蓥蓦蔷蔹蔺蕲薮藓虏虑虚虫虽虾虿蚀蚁蚂蚕蚝蚬蛊蛎蛏蛮蛰蛱蛲蛳蛴蜕蜗蜡蝇蝈蝉蝎蝼蝾螀衅衔补表衬衮袄袜袭装裆裢裣裤裥褛褴见观规觅视觇览觉觊觋觌觎觏觐觑角触誉计订讣认讥讨让讪讫训议讯记讲讳讴讵讶讷许讹论讼讽设访诀证诂诃评诅识诈诉诊词诎诏译诒诔试诗诘诙诚诛诜话诞诟诠诡询诣诤该详诧诨诩诫诬语误诰诱诲诳说诵请诸诹诺读诼诽课诿谀谁谂调谄谅谆谈谊谋谍谎谏谐谑谒谓谔谕谖谗谘谙谚谛谜谝谟谠谡谢谣谤谥谦谧谨谩谪谫谬谭谯谰谱谲谳谴谵谶谷豮贝贞负贡财责贤败账货质贩贪贫贬购贮贯贰贱贲贳贴贵贷贸费贺贻贼贽贾贿赀赁赂赃资赅赆赇赈赉赊赋赌赍赎赏赐赓赔赕赖赘赚赛赜赝赞赠赡赢赣赵赶趋趱跃跄跞践跷跸跹跻踌踪踬踯蹑蹒蹰蹿躏躜躯车轧轨轩轪轫转轭轮软轰轱轲轳轴轵轶轷轸轹轺轻轼载轾轿辀辁辂较辄辅辆辇辈辉辊辋辍辎辏辐辑辒输辔辕辖辗辘辙辚辞辩辫边辽达迁过迈运还这进远违连迟迩迳迹适选逊递逻遗遥邓邝邬邮邻郑郓郦郧郸酝酱酽酾酿释里鉴銮钆钇针钉钊钋钌钍钎钏钐钒钓钔钕钗钙钛钜钝钞钟钠钡钢钣钥钦钧钨钩钪钫钬钭钮钯钰钱钲钳钴钵钶钷钸钹钺钻钼钽钾铀铁铂铃铄铅铆铈铉铊铋铌铍铎铏铐铑铒铕铖铗铘铙铚铛铜铝铞铟铠铡铢铣铤铥铦铨铩铪铫铬铭铮铯铰铱铲铳铴铵银铷铸铺链铿销锁锂锃锅锆锇锈锉锋锌锐锑锒锓锔锕锖锗错锚锛锞锟锡锢锣锤锥锦锨锩锪锫锬锭键锯锰锱锲锴锵锶锷锸锹锺锻锼锾镀镁镂镃镄镅镆镇镉镊镌镍镎镏镐镑镒镓镔镕镖镗镘镙镚镛镜镝镞镟镡镢镣镤镥镦镧镨镩镪镫镬镭镮镯镰镱镲镳镴镵长门闩闪闫闭问闯闰闱闲闳间闵闶闷闸闹闺闻闼闽闾阀阁阂阃阄阅阆阈阉阊阋阌阍阎阏阐阑阒阔阕阖阗阙阚队阳阴阵阶际陆陇陈陉陕陧陨险随隐隶难雏雠雳雾霁霉霭靓静面鞑鞒鞯韦韧韩韪韬页顶顷项顺须顽顾顿颀颁颂预颅领颇颈颉颊颌颍颏颐频颓颔颖颗题颚颛颜额颞颟颠颡颢颤颥颦风飏飐飒飓飔飕飖飘飙飞饥饧饨饩饪饫饬饭饮饯饰饱饲饴饵饶饷饺饼饿馀馁馄馅馆馈馊馋馍馏馐馒馓馔馕马驭驮驯驰驱驳驴驵驶驷驸驹驺驻驼驽驾驿骀骁骂骄骅骆骇骈骊骋验骏骐骑骒骓骔骕骖骗骘骚骛骜骝骞骟骠骡骢骣骤骥骦骧鱼鲁鲂鲅鲆鲇鲈鲋鲍鲎鲐鲑鲒鲔鲕鲚鲛鲜鲞鲟鲠鲡鲢鲣鲥鲦鲧鲨鲩鲫鲮鲰鲱鲲鲳鲴鲵鲶鲷鲸鲺鲻鲼鲽鳀鳁鳃鳄鳅鳆鳇鳌鳍鳎鳏鳐鳓鳔鳕鳖鳗鳘鳙鳜鳝鳞鳟鳢鸟鸡鸢鸣鸥鸦鸧鸨鸩鸪鸫鸬鸭鸯鸱鸲鸳鸵鸶鸷鸸鸹鸺鸽鸾鸿鹁鹂鹃鹄鹅鹆鹇鹈鹉鹊鹋鹌鹎鹏鹑鹕鹗鹘鹚鹛鹜鹞鹣鹤鹦鹧鹨鹩鹪鹫鹬鹭鹰鹱鹳鹾麦黄黉黩齐齑齿龄龀龃龅龆龇龈龉龊龋龌龙龚龛龟"
)


def candidate_files():
    archive = BASE / "資料轉換" / "input" / "archive"
    files = list(archive.glob("*.xlsx"))
    for p in BASE.glob("*.xlsx"):
        n = p.name.upper()
        if any(k in n for k in ("BOM", "TINY", "M90", "SIDECAR", "料號_")):
            files.append(p)
    return sorted(set(files))


def cell_text(v):
    if v is None:
        return ""
    return str(v)


def main():
    occurrences = defaultdict(list)
    audited_fields = 0
    audited_sheets = 0
    audited_files = 0
    mapped = set(cms.SIMP_TO_TRAD)

    for path in candidate_files():
        try:
            wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
        except Exception as exc:
            print(f"SKIP {path}: {exc}")
            continue
        file_used = False
        for ws in wb.worksheets:
            if ws.title in cms.EXCLUDED_SHEETS:
                continue
            if ws.title != "BOM格式" and not cms.is_green_tab(wb, ws):
                continue
            matrix = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]
            meta = cms.detect_sheet_metadata(matrix)
            cols = meta["cols"]
            target_cols = [(role, cols.get(role, -1)) for role in ("zf", "zh", "spec")]
            target_cols = [(r, c) for r, c in target_cols if c is not None and c >= 0]
            if not target_cols:
                continue
            audited_sheets += 1
            file_used = True
            for ridx in range(meta["header_row_idx"] + 1, len(matrix)):
                for role, cidx in target_cols:
                    if cidx >= len(matrix[ridx]):
                        continue
                    text = cell_text(matrix[ridx][cidx])
                    if not text or text.startswith("="):
                        continue
                    audited_fields += 1
                    for ch in set(text):
                        if ch in SIMPLIFIED_ONLY and ch not in mapped:
                            entry = {
                                "file": path.name,
                                "sheet": ws.title,
                                "cell": f"{openpyxl.utils.get_column_letter(cidx+1)}{ridx+1}",
                                "field": role,
                                "text": text,
                            }
                            if entry not in occurrences[ch] and len(occurrences[ch]) < 12:
                                occurrences[ch].append(entry)
        if file_used:
            audited_files += 1
        wb.close()

    result = {
        "audited_files": audited_files,
        "audited_sheets": audited_sheets,
        "audited_nonblank_business_cells": audited_fields,
        "missing_candidates": dict(sorted(occurrences.items())),
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
